import RPi.GPIO as GPIO
from smbus2 import SMBus
import time
import os, sys, serial, time
import numpy as np
import struct,binascii
from threading import Thread
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
import matplotlib as mpl
import matplotlib.style as mplstyle
import pandas as pd
from openpyxl import load_workbook as lw
from datetime import datetime

#better the plot performance 
mpl.rcParams['path.simplify_threshold']=1.0
mpl.rcParams['path.simplify'] = True
mpl.rcParams['agg.path.chunksize'] = 10000
mplstyle.use('fast')

i2c = SMBus(1)
Slave_address = 0x0A
I2c_Param = [0xAA, 0x09, 0x04, 0x96, 0x01, 0x01, 0x00, 0x00, 0x00, 0x9C, 0x00]
I2c_ADC = [0xAA, 0x09, 0x04, 0x98, 0x01, 0x01, 0x00, 0x00, 0x00, 0x9E, 0x00]
I2c_WAVE = [0xAA, 0x09, 0x04, 0x97, 0x01, 0x01, 0x00, 0x00, 0x00, 0x9D, 0x00]
I2c_WAVE_STOP = [0xAA, 0x09, 0x04, 0x97, 0x01, 0x00, 0x00, 0x00, 0x00, 0x9C, 0x00]
param = [
    [0xAA, 0x09, 0x04, 0x80, 0x01, 0x45, 0x01, 0x00, 0x00, 0xCB, 0x00],
    [0xAA, 0x09, 0x04, 0x81, 0x01, 0x08, 0x00, 0x00, 0x00, 0x8E, 0x00],
    [0xAA, 0x09, 0x04, 0x82, 0x01, 0x40, 0x1F, 0x00, 0x00, 0xE6, 0x00],
    [0xAA, 0x09, 0x04, 0x83, 0x01, 0xFA, 0x00, 0x00, 0x00, 0x82, 0x01],
    [0xAA, 0x09, 0x04, 0x84, 0x01, 0x1F, 0x00, 0x00, 0x00, 0xA8, 0x00],
    [0xAA, 0x09, 0x04, 0x85, 0x01, 0x00, 0x00, 0x74, 0x42, 0x40, 0x01],
    [0xAA, 0x09, 0x04, 0x86, 0x01, 0x40, 0x1F, 0x00, 0x00, 0xEA, 0x00],
    [0xAA, 0x09, 0x04, 0x87, 0x01, 0xC8, 0x00, 0x00, 0x00, 0x54, 0x01],
    [0xAA, 0x09, 0x04, 0x88, 0x01, 0xD0, 0x07, 0x00, 0x00, 0x64, 0x01],
    [0xAA, 0x09, 0x04, 0x89, 0x01, 0x28, 0x00, 0x00, 0x00, 0xB6, 0x00],
    [0xAA, 0x09, 0x04, 0x8A, 0x01, 0x00, 0x00, 0x00, 0x00, 0x8F, 0x00],
    [0xAA, 0x09, 0x04, 0x8B, 0x01, 0x00, 0x00, 0x00, 0x00, 0x90, 0x00],
    [0xAA, 0x09, 0x04, 0x8C, 0x01, 0x64, 0x00, 0x00, 0x00, 0xF5, 0x00],
    [0xAA, 0x09, 0x04, 0x8D, 0x01, 0x03, 0x00, 0x00, 0x00, 0x95, 0x00],
    [0xAA, 0x09, 0x04, 0x8E, 0x01, 0x32, 0x00, 0x00, 0x00, 0xC5, 0x00],
    [0xAA, 0x09, 0x04, 0x8F, 0x01, 0x40, 0x0D, 0x03, 0x00, 0xE4, 0x00],
    [0xAA, 0x09, 0x04, 0x90, 0x01, 0x20, 0x4E, 0x00, 0x00, 0x03, 0x01],
    [0xAA, 0x09, 0x04, 0x91, 0x01, 0x10, 0x27, 0x00, 0x00, 0xCD, 0x00],
    [0xAA, 0x09, 0x04, 0x92, 0x01, 0x80, 0x01, 0x00, 0x00, 0x18, 0x01],
    [0xAA, 0x09, 0x04, 0x93, 0x01, 0x0A, 0x00, 0x00, 0x00, 0xA2, 0x00],
    [0xAA, 0x09, 0x04, 0x94, 0x01, 0x00, 0x00, 0x00, 0x00, 0x99, 0x00],
    [0xAA, 0x09, 0x04, 0x9E, 0x01, 0x40, 0x0D, 0x03, 0x00, 0xF3, 0x00],
    [0xAA, 0x09, 0x04, 0x9F, 0x01, 0x78, 0x00, 0x00, 0x00, 0x1C, 0x01],
    [0xAA, 0x09, 0x04, 0x95, 0x01, 0x01, 0x00, 0x00, 0x00, 0x9B, 0x00],
    [0xAA, 0x09, 0x04, 0x9B, 0x01, 0xC8, 0x00, 0xF0, 0x00, 0x58, 0x02],
    [0xAA, 0x09, 0x04, 0x9C, 0x01, 0x00, 0x00, 0x00, 0x00, 0xA1, 0x00],
    [0xAA, 0x09, 0x04, 0xA0, 0x01, 0x00, 0x00, 0x00, 0x00, 0xA5, 0x00]
    ]
DTOF_buffer = []
AbsTOF_UPS_buffer = []
AbsTOF_DNS_buffer = []
VOL_buffer = []
VOL_mean_buffer = []
Standard_VOL_buffer = []
Standard_VOL_mean_buffer = []
Standard_T_buffer = []
Standard_VEL_buffer = []
Digit_T_buffer = []
Digit_P_buffer = []
Accu_buffer = []
UPS_zero_buffer = []
DNS_zero_buffer = []
sum_vol_keti = 0
auto_save = 0

wave_stop_flag = False
wave_start_flag = False

class MyWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()
 
        self.setLayout(self.layout)
        self.setGeometry(50, 50, 1800, 1100)
 
    def initUI(self):
        self.fig = plt.Figure()
        
        self.graph = FigureCanvas(self.fig)
        

        ###ADC
        self.Button_ADC = QPushButton()
        self.Button_ADC.setText("ADC_Capture")
        self.Button_ADC.clicked.connect(self.ADC_capture)
        
        self.Lbl_ADC_max_UPS = QLabel()
        self.Lbl_ADC_max_UPS.setText("Lbl_ADC_max_UPS")
        self.Lbl_ADC_max_UPS.setAlignment(Qt.AlignCenter)
        self.Lbl_ADC_max_DNS = QLabel()
        self.Lbl_ADC_max_DNS.setText("Lbl_ADC_max_DNS")
        self.Lbl_ADC_max_DNS.setAlignment(Qt.AlignCenter)

        ADC_info_layout = QHBoxLayout()
        ADC_info_layout.addWidget(self.Lbl_ADC_max_UPS)
        ADC_info_layout.addWidget(self.Lbl_ADC_max_DNS)
        ADC_info_layout.addWidget(self.Button_ADC)
        
        ADC_layout = QVBoxLayout()
        ADC_layout.addLayout(ADC_info_layout)
        #ADC_layout.addWidget(self.ADC)
        
        

        ###DTOF
        self.Button_wave = QPushButton()
        self.Button_wave.setText("Wave_Start")
        self.Button_wave.clicked.connect(self.Wave_capture)
        
        self.Lbl_DTOF_mean = QLabel()
        self.Lbl_DTOF_mean.setText("Lbl_DTOF_mean")
        self.Lbl_DTOF_mean.setAlignment(Qt.AlignCenter)

        wave_info_layout = QHBoxLayout()
        wave_info_layout.addWidget(self.Lbl_DTOF_mean)
        wave_info_layout.addWidget(self.Button_wave)
        
        DTOF_layout = QVBoxLayout()
        DTOF_layout.addLayout(wave_info_layout)
        #DTOF_layout.addWidget(self.DTOF)
        

        ###AbsTOF
        self.Lbl_AbsTOF_UPS_mean = QLabel()
        self.Lbl_AbsTOF_UPS_mean.setText("Lbl_AbsTOF_UPS_mean")
        self.Lbl_AbsTOF_UPS_mean.setAlignment(Qt.AlignCenter)
        self.Lbl_AbsTOF_DNS_mean = QLabel()
        self.Lbl_AbsTOF_DNS_mean.setText("Lbl_AbsTOF_DNS_mean")
        self.Lbl_AbsTOF_DNS_mean.setAlignment(Qt.AlignCenter)

        AbsTOF_info_layout = QHBoxLayout()
        AbsTOF_info_layout.addWidget(self.Lbl_AbsTOF_UPS_mean)
        AbsTOF_info_layout.addWidget(self.Lbl_AbsTOF_DNS_mean)
        
        AbsTOF_layout = QVBoxLayout()
        AbsTOF_layout.addLayout(AbsTOF_info_layout)
        #AbsTOF_layout.addWidget(self.ABS)

        ###Volume
        self.Lbl_Volume_mean = QLabel()
        self.Lbl_Volume_mean.setText("Lbl_Volume_mean")
        self.Lbl_Volume_mean.setAlignment(Qt.AlignCenter)

        self.Lbl_Standard_Volume_mean = QLabel()
        self.Lbl_Standard_Volume_mean.setText("Lbl_Volume_mean")
        self.Lbl_Standard_Volume_mean.setAlignment(Qt.AlignCenter)

        self.Lbl_Accuracy = QLabel()
        self.Lbl_Accuracy.setText("Accuracy")
        self.Lbl_Accuracy.setAlignment(Qt.AlignCenter)


        Volume_info_layout = QHBoxLayout()
        Volume_info_layout.addWidget(self.Lbl_Volume_mean)
        Volume_info_layout.addWidget(self.Lbl_Standard_Volume_mean)
        Volume_info_layout.addWidget(self.Lbl_Accuracy)

        
        Volume_layout = QVBoxLayout()
        Volume_layout.addLayout(Volume_info_layout)
        #Volume_layout.addWidget(self.VOL)

        ###VT
        Button_save_wave = QPushButton()
        Button_save_wave.setText("Wave_Save")
        Button_save_wave.clicked.connect(self.Wave_save)

        Button_clear_wave = QPushButton()
        Button_clear_wave.setText("Wave_Clear")
        Button_clear_wave.clicked.connect(self.Wave_clear)

        Button_SS_layout = QHBoxLayout()
        Button_SS_layout.addWidget(Button_save_wave)
        Button_SS_layout.addWidget(Button_clear_wave)

        VT_layout = QVBoxLayout()
        VT_layout.addLayout(Button_SS_layout)
        #VT_layout.addWidget(self.V_T)

        ###VP
        self.Lbl_T_Pres_mean = QLabel()
        self.Lbl_T_Pres_mean.setText("Lbl_Temperature_Pressure_mean")
        self.Lbl_T_Pres_mean.setAlignment(Qt.AlignCenter)
        
        VP_layout = QVBoxLayout()
        VP_layout.addWidget(self.Lbl_T_Pres_mean)
        #VP_layout.addWidget(self.V_P)

        ###Param
        Lbl_Transmit_freq = QLabel()
        Lbl_Transmit_freq.setText("Transmit_Freq")
        Lbl_Transmit_freq.setAlignment(Qt.AlignCenter)
        Lbl_ADC_Gap = QLabel()
        Lbl_ADC_Gap.setText("ADC_Start_Gap")
        Lbl_ADC_Gap.setAlignment(Qt.AlignCenter)
        Lbl_Num_Pulses = QLabel()
        Lbl_Num_Pulses.setText("Num of Pulses")
        Lbl_Num_Pulses.setAlignment(Qt.AlignCenter)
        Lbl_UPS2DNS = QLabel()
        Lbl_UPS2DNS.setText("UPS_DNS_Gap")
        Lbl_UPS2DNS.setAlignment(Qt.AlignCenter)
        Lbl_UPS2UPS = QLabel()
        Lbl_UPS2UPS.setText("UPS_UPS_Gap")
        Lbl_UPS2UPS.setAlignment(Qt.AlignCenter)
        self.Lbl_Gain = QLabel()
        self.Lbl_Gain.setText("Gain : 5.2")
        self.Lbl_Gain.setAlignment(Qt.AlignCenter)
        Lbl_Meter_const = QLabel()
        Lbl_Meter_const.setText("Meter Constant")
        Lbl_Meter_const.setAlignment(Qt.AlignCenter)
        Lbl_USSXT = QLabel()
        Lbl_USSXT.setText("USSXT")
        Lbl_USSXT.setAlignment(Qt.AlignCenter)
        Lbl_Sig_freq = QLabel()
        Lbl_Sig_freq.setText("Signal Freq")
        Lbl_Sig_freq.setAlignment(Qt.AlignCenter)
        Lbl_DTOF_offset = QLabel()
        Lbl_DTOF_offset.setText("DTOF Offset")
        Lbl_DTOF_offset.setAlignment(Qt.AlignCenter)
        Lbl_AbsTOF_offset = QLabel()
        Lbl_AbsTOF_offset.setText("AbsTOF offset")
        Lbl_AbsTOF_offset.setAlignment(Qt.AlignCenter)
        Lbl_cap_duration = QLabel()
        Lbl_cap_duration.setText("Capture Duration")
        Lbl_cap_duration.setAlignment(Qt.AlignCenter)
        

        self.txt_Transmit_freq = QLineEdit()
        self.txt_Transmit_freq.setText("200")
        self.txt_ADC_Gap = QLineEdit()
        self.txt_ADC_Gap.setText("325")
        self.txt_Num_Pulses = QLineEdit()
        self.txt_Num_Pulses.setText("8")
        self.txt_UPS2DNS = QLineEdit()
        self.txt_UPS2DNS.setText("8000")
        self.txt_UPS2UPS = QLineEdit()
        self.txt_UPS2UPS.setText("250")
        self.Combo_Gain = QSpinBox()
        self.Combo_Gain.setFixedSize(200,32)
        self.Combo_Gain.setValue(31)
        self.Combo_Gain.setSingleStep(1)
        self.Combo_Gain.setMinimum(17)
        self.Combo_Gain.setMaximum(63)
        self.Combo_Gain.valueChanged.connect(self.GainChanged)
        self.txt_Meter_const = QLineEdit()
        self.txt_Meter_const.setText("61")
        self.txt_USSXT = QLineEdit()
        self.txt_USSXT.setText("8000")
        self.txt_Sig_freq = QLineEdit()
        self.txt_Sig_freq.setText("2000")
        self.txt_DTOF_offset = QLineEdit()
        self.txt_DTOF_offset.setText("0")
        self.txt_AbsTOF_offset = QLineEdit()
        self.txt_AbsTOF_offset.setText("0")
        self.txt_cap_duration = QLineEdit()
        self.txt_cap_duration.setText("100")
        
        self.Button_Params = QPushButton()
        self.Button_Params.setText("Set_Params")
        self.Button_Params.clicked.connect(self.Set_Params)

        Param_Label_H_layout = QHBoxLayout()
        Param_Label_H_layout.addWidget(Lbl_Transmit_freq)
        Param_Label_H_layout.addWidget(Lbl_ADC_Gap)
        Param_Label_H_layout.addWidget(Lbl_Num_Pulses)
        Param_Label_H_layout.addWidget(Lbl_UPS2DNS)
        Param_Label_H_layout.addWidget(Lbl_UPS2UPS)
        Param_Label_H_layout.addWidget(self.Lbl_Gain)
        Param_Label_H_layout.addWidget(Lbl_Meter_const)

        Param_Label_L_layout = QHBoxLayout()
        Param_Label_L_layout.addWidget(Lbl_USSXT)
        Param_Label_L_layout.addWidget(Lbl_Sig_freq)
        Param_Label_L_layout.addWidget(Lbl_DTOF_offset)
        Param_Label_L_layout.addWidget(Lbl_AbsTOF_offset)
        Param_Label_L_layout.addWidget(Lbl_cap_duration)
        
        Param_txt_H_layout = QHBoxLayout()
        Param_txt_H_layout.addWidget(self.txt_Transmit_freq)
        Param_txt_H_layout.addWidget(self.txt_ADC_Gap)
        Param_txt_H_layout.addWidget(self.txt_Num_Pulses)
        Param_txt_H_layout.addWidget(self.txt_UPS2DNS)
        Param_txt_H_layout.addWidget(self.txt_UPS2UPS)
        Param_txt_H_layout.addWidget(self.Combo_Gain)
        Param_txt_H_layout.addWidget(self.txt_Meter_const)
        
        Param_txt_L_layout = QHBoxLayout()
        Param_txt_L_layout.addWidget(self.txt_USSXT)
        Param_txt_L_layout.addWidget(self.txt_Sig_freq)
        Param_txt_L_layout.addWidget(self.txt_DTOF_offset)
        Param_txt_L_layout.addWidget(self.txt_AbsTOF_offset)
        Param_txt_L_layout.addWidget(self.txt_cap_duration)

        Param_layout = QVBoxLayout()
        Param_layout.addLayout(Param_Label_H_layout)
        Param_layout.addLayout(Param_txt_H_layout)
        Param_layout.addLayout(Param_Label_L_layout)
        Param_layout.addLayout(Param_txt_L_layout)
        Param_layout.addWidget(self.Button_Params)

        wave_H_layout = QHBoxLayout()
        wave_H_layout.addLayout(ADC_layout)
        wave_H_layout.addLayout(DTOF_layout)
        wave_H_layout.addLayout(VT_layout)
        
        wave_L_layout = QHBoxLayout()
        wave_L_layout.addLayout(AbsTOF_layout)
        wave_L_layout.addLayout(Volume_layout)
        wave_L_layout.addLayout(VP_layout)
        
        layout = QVBoxLayout()
        layout.addLayout(Param_layout)
        layout.addLayout(wave_H_layout)
        layout.addLayout(wave_L_layout)
        layout.addWidget(self.graph)
        
        self.layout = layout


    def GainChanged(self):
        Gain_LUT = [-6.5,-5.5,-4.6,-4.1,-3.3,-2.3,-1.4,-0.8,0.1,1.0,1.9,2.6,3.5,
                    4.4,5.2,6.0,6.8,7.7,8.7,9.0,9.8,10.7,11.7,12.2,13.0,13.9,14.9,
                    15.5,16.3,17.2,18.2,18.8,19.6,20.5,21.5,22.0,22.8,23.6,24.6,
                    25.0,25.8,26.7,27.7,28.1,28.9,29.8,30.8]
        
        self.Lbl_Gain.setText("Gain : " + str(Gain_LUT[self.Combo_Gain.value()-17]))

    def Wave_capture(self):
        if (self.Button_wave.text() == "Wave_Start"):
            ###self.fig.clear()
            ###self.fig.add_subplot(111)
            ###self.ADC.draw()
            self.Button_wave.setText("Wave_Stop")
            Wave_Start = Thread(target=self.Wave_capture_start)
            Wave_Start.start()
        else:
            self.Button_wave.setText("Wave_Start")
            self.Wave_stop()

    def Wave_capture_start(self):
        global DTOF_buffer
        global AbsTOF_UPS_buffer
        global AbsTOF_DNS_buffer
        global VOL_buffer
        global VOL_mean_buffer
        global Standard_VOL_mean_buffer
        global Standard_VOL_buffer
        global Standard_T_buffer
        global Standard_VEL_buffer
        global wave_stop_flag
        global wave_start_flag
        global sum_vol_keti
        global Digit_T_buffer
        global Digit_P_buffer
        global Accu_buffer
        global UPS_zero_buffer
        global DNS_zero_buffer
        wave_start_flag = True

        GPIO.setmode(GPIO.BCM)
        GPIO.setup(17,GPIO.IN,pull_up_down=GPIO.PUD_UP)
        
        i2c.write_i2c_block_data(Slave_address,0x55,I2c_WAVE)
        DTOF_data = []
        while(1):
            if(wave_stop_flag):
                wave_stop_flag=False
                break
            else:
                P = GPIO.wait_for_edge(17,GPIO.FALLING,timeout=2000)
            if P is None:
                #print("Timeout")
                #break
                pass
            else:
                raw_data = i2c.read_i2c_block_data(Slave_address,0,20)
                raw_list = []
                raw_buffer = ""
                for i in range(5,1,-1):
                    raw_buffer = raw_buffer + format(raw_data[i],'x').zfill(2)
                raw_hex = binascii.unhexlify(raw_buffer)
                raw_float = struct.unpack('!f',raw_hex)[0]
                raw_list.append(raw_float)
                raw_buffer = ""
                for i in range(9,5,-1):
                    raw_buffer = raw_buffer + format(raw_data[i],'x').zfill(2)
                raw_hex = binascii.unhexlify(raw_buffer)
                raw_float = struct.unpack('!f',raw_hex)[0]
                raw_list.append(raw_float)
                raw_buffer = ""
                for i in range(13,9,-1):
                    raw_buffer = raw_buffer + format(raw_data[i],'x').zfill(2)
                raw_hex = binascii.unhexlify(raw_buffer)
                raw_float = struct.unpack('!f',raw_hex)[0]
                raw_list.append(raw_float)
                raw_buffer = ""
                for i in range(17,13,-1):
                    raw_buffer = raw_buffer + format(raw_data[i],'x').zfill(2)
                raw_hex = binascii.unhexlify(raw_buffer)
                raw_float = struct.unpack('!f',raw_hex)[0]
                raw_list.append(raw_float)
                DTOF_buffer.append(raw_list[0])
                AbsTOF_UPS_buffer.append(raw_list[1])
                AbsTOF_DNS_buffer.append(raw_list[2])
                VOL_buffer.append(raw_list[3])

                #if(raw_list[3]>0.667):
                    #sum_vol_keti = sum_vol_keti + raw_list[3]

                ###UPS_zero = ((raw_list[1]*1000000000)+40000-(int(self.txt_ADC_Gap.text())*1000))/(1000000/int(self.txt_Sig_freq.text()))
                ###UPS_zero_buffer.append(UPS_zero)
                ###DNS_zero = ((raw_list[2]*1000000000)+40000-(int(self.txt_ADC_Gap.text())*1000))/(1000000/int(self.txt_Sig_freq.text()))
                ###DNS_zero_buffer.append(DNS_zero)
                
                ###self.Lbl_ADC_max_UPS.setText("UPS_Zero = {0:.3f}".format(UPS_zero))
                ###self.Lbl_ADC_max_DNS.setText("DNS_Zero = {0:.3f}".format(DNS_zero))

                #GN = Graph data length
                GN = 100

                ###self.wave_zeroX(UPS_zero_buffer[-GN:],DNS_zero_buffer[-GN:])

                #N_point_moving_avg
                N = 3 #3_Point
                DTOF_mean = np.mean(np.array(DTOF_buffer[-N:]))
                AbsTOF_UPS_mean = np.mean(np.array(AbsTOF_UPS_buffer[-N:]))
                AbsTOF_DNS_mean = np.mean(np.array(AbsTOF_DNS_buffer[-N:]))
                Volume_mean = np.mean(np.array(VOL_buffer[-N:]))

                VOL_mean_buffer.append(Volume_mean)

                self.Lbl_DTOF_mean.setText("DTOF = {0:.3f} [ns]".format(DTOF_mean*1000000000))
                self.Lbl_AbsTOF_UPS_mean.setText("ATOF_U = {0:.3f} [us]".format(AbsTOF_UPS_mean*1000000))
                self.Lbl_AbsTOF_DNS_mean.setText("ATOF_D = {0:.3f} [us]".format(AbsTOF_DNS_mean*1000000))
                #self.Lbl_Volume_mean.setText("Vol = {0:.3f} [l/min], {1:.3f} [m3/h], Sum = {2:.3f} [l/min]".format(Volume_mean,Volume_mean*60/1000,sum_vol_keti))
                self.Lbl_Volume_mean.setText("Vol = {0:.3f} [l/min], {1:.3f} [m3/h]".format(Volume_mean,Volume_mean*60/1000))

                #self.wave_dtof(DTOF_buffer[-GN:])
                #self.wave_abstof(AbsTOF_UPS_buffer[-GN:],AbsTOF_DNS_buffer[-GN:])
                #self.wave_vol(VOL_buffer[-GN:],VOL_mean_buffer[-GN:])

                #standard
                ser = serial.Serial('/dev/ttyUSB0', 9600, 8, timeout = 0.1)
                ser.write(bytes(bytearray([0x01,0x03,0x00,0x00,0x00,0x14,0x45,0xC5])))
                large = ser.read(50)
                
                float_current_flow = 0
                float_current_flowvel = 0
                try:
                    result_large = ''.join(format(x,'02x') for x in large)
                    temperature_large = result_large[10:14]+result_large[6:10]
                    raw_hex = binascii.unhexlify(temperature_large)
                    float_T_large = struct.unpack('!f',raw_hex)[0]
                    current_flowrate_large = result_large[34:38] + result_large[30:34]
                    raw_hex = binascii.unhexlify(current_flowrate_large)
                    float_current_flow = struct.unpack('!f',raw_hex)[0]
                    Standard_T_buffer.append(float_T_large)
                    current_flowvel_large = result_large[26:30] + result_large[22:26]
                    raw_hex = binascii.unhexlify(current_flowvel_large)
                    float_current_flowvel = struct.unpack('!f',raw_hex)[0]
                    Standard_VEL_buffer.append(float_current_flowvel)
                except:
                    if(len(Standard_VOL_buffer)!=0):
                        Standard_T_buffer.append(Standard_VOL_buffer[-1])
                        float_current_flow = Standard_VOL_buffer[-1]
            
                if(float_current_flow>0):
                    Standard_VOL_buffer.append(float_current_flow)
                else:
                    ser.write(bytes(bytearray([0x02,0x04,0x00,0x00,0x00,0x03,0xB0,0x38])))
                    small = ser.read(50)
                    try:
                        result_small = ''.join(format(x,'02x') for x in small)
                        current_flowrate_small = result_small[14:18]
                        int_flow = int(current_flowrate_small,16)*0.01
                        Standard_VOL_buffer.append(int_flow)
                    except:
                        Standard_VOL_buffer.append(Standard_VOL_buffer[-1])

                Standard_VOL_mean = np.mean(np.array(Standard_VOL_buffer[-N:]))
                Standard_VOL_mean_buffer.append(Standard_VOL_mean)
                self.Lbl_Standard_Volume_mean.setText("Std_Vol = {0:.3f} [l/min]".format(Standard_VOL_mean))
                
                

                ser.write(bytes(bytearray([0x03,0x04,0x00,0x00,0x00,0x01,0x30,0x28])))
                T_buff = ser.read(10)
                result_T = ''.join(format(x,'02x') for x in T_buff)
                val_T = 0
                try:
                    val_T = int(result_T[8:10],16)*0.1
                    if(val_T<10):
                        val_T = Digit_T_buffer[-1]
                    if((len(Digit_T_buffer)!=0) and abs(val_T - Digit_T_buffer[-1])>0.5 ):
                        val_T = Digit_T_buffer[-1]
                    Digit_T_buffer.append(val_T)
                except:
                    if(len(Digit_T_buffer)!=0):
                        Digit_T_buffer.append(Digit_T_buffer[-1])
                
                #self.wave_T(Digit_T_buffer[-GN:],VOL_mean_buffer[-GN:])

                ser.write(bytes(bytearray([0x04,0x04,0x00,0x00,0x00,0x01,0x31,0x9F])))
                P_buff = ser.read(10)
                result_P = ''.join(format(x,'02x') for x in P_buff)
                val_P = 0
                try:
                    val_P = int(result_P[7:10],16)
                    if((len(Digit_P_buffer)!=0) and abs(val_P - Digit_P_buffer[-1])>20 ):
                        val_P = Digit_P_buffer[-1]
                    Digit_P_buffer.append(val_P)
                except:
                    if(len(Digit_P_buffer)!=0):
                        Digit_P_buffer.append(Digit_P_buffer[-1])
                #self.wave_P(Digit_P_buffer[-GN:],VOL_mean_buffer[-GN:])

                val_Accu = Volume_mean/Standard_VOL_mean
                self.Lbl_Accuracy.setText("Accuracy = {0:.3f}%".format(val_Accu*100))
                Accu_buffer.append(val_Accu)
                self.Lbl_T_Pres_mean.setText("Stan_T = {0:.3f}, Stan_VEL = {3:.3f},  Thermo  = {1:.1f},   Pressure = {2:}".format(Standard_T_buffer[-1],Digit_T_buffer[-1],Digit_P_buffer[-1],Standard_VEL_buffer[-1]))

                #self.Lbl_ADC_max_UPS.setText("UPS_max = [{0}, {1}]".format(y.index(max(y)),max(y)))
                #self.Lbl_ADC_max_DNS.setText("DNS_max = [{0}, {1}]".format(y2.index(max(y2)),max(y2)))

                self.wave_graph(DTOF_buffer[-GN:],AbsTOF_UPS_buffer[-GN:],AbsTOF_DNS_buffer[-GN:],VOL_buffer[-GN:],VOL_mean_buffer[-GN:],Digit_T_buffer[-GN:],Digit_P_buffer[-GN:])

                #global auto_save
                #auto_save = auto_save + 1
                #if(auto_save == 100):
                #    auto_save = 0
                #    self.Wave_save()
                #    time.sleep(2)
                #    self.Wave_clear()
                #    time.sleep(3)

    def wave_graph(self,DTOF,UPS,DNS,VOL,VOL_mean,T,P):
        self.fig.clear()
        g1 = self.fig.add_subplot(231)
        g1.plot(DTOF,'r')

        g2 = self.fig.add_subplot(232)
        g2.plot(UPS,'r')
        g2.plot(DNS,'b')

        g3 = self.fig.add_subplot(233)
        g3.plot(VOL,'k')
        g3.plot(VOL_mean,'r--')

        g4 = self.fig.add_subplot(234)
        g4.plot(T,VOL_mean,'ro')

        g5 = self.fig.add_subplot(235)
        g5.plot(P,VOL_mean,'bo')

        self.graph.draw()
                    

    def wave_zeroX(self,UPS,DNS):
        self.fig.clear()
        ax = self.fig.add_subplot(111)
        ax.plot(UPS,'r')
        ax.plot(DNS,'b')
        
        self.ADC.draw()

    def wave_dtof(self,buff):
        self.fig2.clear()
        dtof = self.fig2.add_subplot(111)
        dtof.plot(buff,'r')
        self.DTOF.draw()

    def wave_abstof(self,ups,dns):
        self.fig3.clear()
        abstof = self.fig3.add_subplot(111)
        abstof.plot(ups,'r')
        abstof.plot(dns,'b')
        self.ABS.draw()

    def wave_vol(self,buff,mean):
        self.fig4.clear()
        vol = self.fig4.add_subplot(111)
        vol.plot(buff,'k')
        vol.plot(mean,'r--')
        self.VOL.draw()

    def wave_T(self,x,y):
        self.fig5.clear()
        T = self.fig5.add_subplot(111)
        T.plot(x,y,'ro')
        self.V_T.draw()

    def wave_P(self,x,y):
        self.fig6.clear()
        P = self.fig6.add_subplot(111)
        P.plot(x,y,'bo')
        self.V_P.draw()

    def Wave_stop(self):
        GPIO.setmode(GPIO.BCM)
        GPIO.setup(17,GPIO.IN,pull_up_down=GPIO.PUD_UP)
        
        i2c.write_i2c_block_data(Slave_address,0x55,I2c_WAVE_STOP)
        GPIO.cleanup()
        global wave_stop_flag
        global wave_start_flag
        wave_start_flag = False
        wave_stop_flag = True

    def Wave_save(self):
        global DTOF_buffer
        global AbsTOF_UPS_buffer
        global AbsTOF_DNS_buffer
        global VOL_buffer
        global VOL_mean_buffer
        global sum_vol_keti
        global Standard_VOL_mean_buffer
        global Standard_VOL_buffer
        global Standard_T_buffer
        global Standard_VEL_buffer
        global Digit_T_buffer
        global Digit_P_buffer
        global Accu_buffer
        global UPS_zero_buffer
        global DNS_zero_buffer


        wb = lw(filename = 'blank.xlsx', read_only=False, data_only=False)
        ws = wb['Sheet1']

        ws['A1'] = "DTOF"
        ws['B1'] = "ABSTOF_UPS"
        ws['C1'] = "ABSTOF_DNS"

        ws['D1'] = "Stan_Temperature"
        ws['E1'] = "Temperature"
        ws['F1'] = "Pressure"

        ws['G1'] = "VOL"
        ws['H1'] = "Stan_VOL"
        ws['I1'] = "VOL_mean"
        ws['J1'] = "Stan_VOL_mean"
        ws['K1'] = "Accurancy"

        ws['L1'] = "Zero_UPS"
        ws['M1'] = "Zero_DNS"

        ws['O1'] = "Stan_Velocity"

        ws['P1'] = "UPS"
        ws['Q1'] = "DNS"
        ws['R1'] = "Fail?"

        for i in range(len(DTOF_buffer)):
            ws['A' + str(i+2)] = DTOF_buffer[i]
        for i in range(len(AbsTOF_UPS_buffer)):
            ws['B' + str(i+2)] = AbsTOF_UPS_buffer[i]
        for i in range(len(AbsTOF_DNS_buffer)):
            ws['C' + str(i+2)] = AbsTOF_DNS_buffer[i]
        for i in range(len(Standard_T_buffer)):
            ws['D' + str(i+2)] = Standard_T_buffer[i]
        for i in range(len(Digit_T_buffer)):
            ws['E' + str(i+2)] = Digit_T_buffer[i]
        for i in range(len(Digit_P_buffer)):
            ws['F' + str(i+2)] = Digit_P_buffer[i]
        for i in range(len(VOL_buffer)):
            ws['G' + str(i+2)] = VOL_buffer[i]
        for i in range(len(Standard_VOL_buffer)):
            ws['H' + str(i+2)] = Standard_VOL_buffer[i]
        for i in range(len(VOL_mean_buffer)):
            ws['I' + str(i+2)] = VOL_mean_buffer[i]
        for i in range(len(Standard_VOL_mean_buffer)):
            ws['J' + str(i+2)] = Standard_VOL_mean_buffer[i]
        for i in range(len(Accu_buffer)):
            ws['K' + str(i+2)] = Accu_buffer[i]
        for i in range(len(UPS_zero_buffer)):
            ws['L' + str(i+2)] = UPS_zero_buffer[i]
        for i in range(len(DNS_zero_buffer)):
            ws['M' + str(i+2)] = DNS_zero_buffer[i]
        for i in range(len(Standard_VEL_buffer)):
            ws['O' + str(i+2)] = Standard_VEL_buffer[i]
        
        
        
        now = datetime.now()
        wb.save("Wave_Data_" + str(now.month) + "_" + str(now.day) + "_" + str(now.hour) + "_" + str(now.minute) + ".xlsx")

    def Wave_clear(self):
        self.fig.clear()
        self.fig2.clear()
        self.fig3.clear()
        self.fig4.clear()
        self.fig5.clear()
        self.fig6.clear()

        global DTOF_buffer
        global AbsTOF_UPS_buffer
        global AbsTOF_DNS_buffer
        global VOL_buffer
        global VOL_mean_buffer
        global sum_vol_keti
        global Standard_VOL_mean_buffer
        global Standard_VOL_buffer
        global Standard_T_buffer
        global Standard_VEL_buffer
        global Digit_T_buffer
        global Digit_P_buffer
        global Accu_buffer
        global UPS_zero_buffer
        global DNS_zero_buffer
        sum_vol_keti = 0

        DTOF_buffer = []
        AbsTOF_UPS_buffer = []
        AbsTOF_DNS_buffer = []
        VOL_buffer = []
        VOL_mean_buffer = []
        Standard_VOL_mean_buffer = []
        Standard_VOL_buffer = []
        Standard_T_buffer = []
        Standard_VEL_buffer = []
        Digit_T_buffer = []
        Digit_P_buffer = []
        Accu_buffer = []
        UPS_zero_buffer = []
        DNS_zero_buffer = []

        self.fig.add_subplot(111)
        self.ADC.draw()

        dtof = self.fig2.add_subplot(111)
        self.DTOF.draw()

        abstof = self.fig3.add_subplot(111)
        self.ABS.draw()

        vol = self.fig4.add_subplot(111)
        self.VOL.draw()

        self.fig5.add_subplot(111)
        self.V_T.draw()

        self.fig6.add_subplot(111)
        self.V_P.draw()

    def Set_Params(self):
        global I2c_Param
        global param
        Param_flag = False
        self.Button_Params.setText("Set_Params")
        
        param[0][5] = int(format(int(self.txt_ADC_Gap.text()),'x').zfill(4)[2:4],16)
        param[0][6] = int(format(int(self.txt_ADC_Gap.text()),'x').zfill(4)[0:2],16)
        param[0][9] = (param[0][3]+param[0][5]+param[0][6]+5)%0x100
        param[0][10] = (param[0][3]+param[0][5]+param[0][6]+5)//0x100
        param[1][5] = int(format(int(self.txt_Num_Pulses.text()),'x').zfill(4)[2:4],16)
        param[1][6] = int(format(int(self.txt_Num_Pulses.text()),'x').zfill(4)[0:2],16)
        param[1][9] = (param[1][3]+param[1][5]+param[1][6]+5)%0x100
        param[1][10] = (param[1][3]+param[1][5]+param[1][6]+5)//0x100
        param[2][5] = int(format(int(self.txt_UPS2DNS.text()),'x').zfill(4)[2:4],16)
        param[2][6] = int(format(int(self.txt_UPS2DNS.text()),'x').zfill(4)[0:2],16)
        param[2][9] = (param[2][3]+param[2][5]+param[2][6]+5)%0x100
        param[2][10] = (param[2][3]+param[2][5]+param[2][6]+5)//0x100
        param[3][5] = int(format(int(self.txt_UPS2UPS.text()),'x').zfill(4)[2:4],16)
        param[3][6] = int(format(int(self.txt_UPS2UPS.text()),'x').zfill(4)[0:2],16)
        param[3][9] = (param[3][3]+param[3][5]+param[3][6]+5)%0x100
        param[3][10] = (param[3][3]+param[3][5]+param[3][6]+5)//0x100
        param[4][5] = int(format(self.Combo_Gain.value(),'x').zfill(4)[2:4],16)
        param[4][6] = int(format(self.Combo_Gain.value(),'x').zfill(4)[0:2],16)
        param[4][9] = (param[4][3]+param[4][5]+param[4][6]+5)%0x100
        param[4][10] = (param[4][3]+param[4][5]+param[4][6]+5)//0x100
        param[5][5] = int(format(struct.unpack('<I',struct.pack('<f',float(self.txt_Meter_const.text())))[0],'x')[6:8],16)
        param[5][6] = int(format(struct.unpack('<I',struct.pack('<f',float(self.txt_Meter_const.text())))[0],'x')[4:6],16)
        param[5][7] = int(format(struct.unpack('<I',struct.pack('<f',float(self.txt_Meter_const.text())))[0],'x')[2:4],16)
        param[5][8] = int(format(struct.unpack('<I',struct.pack('<f',float(self.txt_Meter_const.text())))[0],'x')[0:2],16)
        param[5][9] = (param[5][3]+param[5][5]+param[5][6]+param[5][7]+param[5][8]+5)%0x100
        param[5][10] = (param[5][3]+param[5][5]+param[5][6]+param[5][7]+param[5][8]+5)//0x100
        param[6][5] = int(format(int(self.txt_USSXT.text()),'x').zfill(4)[2:4],16)
        param[6][6] = int(format(int(self.txt_USSXT.text()),'x').zfill(4)[0:2],16)
        param[6][9] = (param[6][3]+param[6][5]+param[6][6]+5)%0x100
        param[6][10] = (param[6][3]+param[6][5]+param[6][6]+5)//0x100
        param[8][5] = int(format(int(self.txt_Sig_freq.text()),'x').zfill(4)[2:4],16)
        param[8][6] = int(format(int(self.txt_Sig_freq.text()),'x').zfill(4)[0:2],16)
        param[8][9] = (param[8][3]+param[8][5]+param[8][6]+5)%0x100
        param[8][10] = (param[8][3]+param[8][5]+param[8][6]+5)//0x100
        param[10][5] = int(format(int(self.txt_DTOF_offset.text()),'x').zfill(4)[2:4],16)
        param[10][6] = int(format(int(self.txt_DTOF_offset.text()),'x').zfill(4)[0:2],16)
        param[10][9] = (param[10][3]+param[10][5]+param[10][6]+5)%0x100
        param[10][10] = (param[10][3]+param[10][5]+param[10][6]+5)//0x100
        param[11][5] = int(format(int(self.txt_AbsTOF_offset.text()),'x').zfill(4)[2:4],16)
        param[11][6] = int(format(int(self.txt_AbsTOF_offset.text()),'x').zfill(4)[0:2],16)
        param[11][9] = (param[11][3]+param[11][5]+param[11][6]+5)%0x100
        param[11][10] = (param[11][3]+param[11][5]+param[11][6]+5)//0x100
        param[12][5] = int(format(int(self.txt_cap_duration.text()),'x').zfill(4)[2:4],16)
        param[12][6] = int(format(int(self.txt_cap_duration.text()),'x').zfill(4)[0:2],16)
        param[12][9] = (param[12][3]+param[12][5]+param[12][6]+5)%0x100
        param[12][10] = (param[12][3]+param[12][5]+param[12][6]+5)//0x100
        param[24][5] = int(format(int(self.txt_Transmit_freq.text()),'x').zfill(4)[2:4],16)
        param[24][6] = int(format(int(self.txt_Transmit_freq.text()),'x').zfill(4)[0:2],16)
        param[24][9] = (param[24][3]+param[24][5]+param[24][6]+param[24][7]+param[24][8]+5)%0x100
        param[24][10] = (param[24][3]+param[24][5]+param[24][6]+param[24][7]+param[24][8]+5)//0x100

        GPIO.setmode(GPIO.BCM)
        GPIO.setup(17,GPIO.IN,pull_up_down=GPIO.PUD_UP)
        
        i2c.write_i2c_block_data(Slave_address,0x55,I2c_Param)

        while(1):
            P = GPIO.wait_for_edge(17,GPIO.FALLING,timeout=4000)
            if P is None:
                print("Timeout")
                break
            else:
                raw_data = i2c.read_i2c_block_data(Slave_address,0,15)
                for i in range(len(param)):
                    if(raw_data[2]==param[i][3]):
                        i2c.write_i2c_block_data(Slave_address,0x55,param[i])
                        break
                    if(raw_data[2]==160):
                        Param_flag = True
        if(Param_flag):
            self.Button_Params.setStyleSheet("background-color:rgb(33,247,21)")
        else:
            self.Button_Params.setStyleSheet("background-color:rgb(247,33,21)")
            self.Button_Params.setText("FAIL")
        #self.txt_Num_Pulses.setText(self.txt_Transmit_freq.text())
        
    def ADC_capture(self):
        self.Button_Params.setStyleSheet("background-color:rgb(255,255,255)")
        self.Button_ADC.setText("Wait...")
        global wave_start_flag
        if(wave_start_flag):
            pass
        else:
            self.fig.clear()

            GPIO.setmode(GPIO.BCM)
            GPIO.setup(17,GPIO.IN,pull_up_down=GPIO.PUD_UP)
            
            i2c.write_i2c_block_data(Slave_address,0x55,I2c_ADC)
            data = []
            while(1):
                P = GPIO.wait_for_edge(17,GPIO.FALLING,timeout=3000)
                if P is None:
                    #print("Timeout")
                    break
                else:
                    raw_data = i2c.read_i2c_block_data(Slave_address,0,32)
                    data.append(raw_data)
            x=[]
            y=[]
            y2=[]
            try:
                for i in range(1,201):
                    x.append(i)
                
                #UPS
                for i in range(1,16):
                    for j in range(1,14):
                        if(data[i][2*j+5]>=128):
                            y.append(-1*((255^data[i][2*j+4])+1+(256*(data[i][2*j+5]^255))))
                        else:
                            y.append(data[i][2*j+4]+(256*data[i][2*j+5]))
                for j in range(1,6):
                    if(data[16][2*j+5]>=128):
                        y.append(-1*((255^data[16][2*j+4])+1+(256*(data[16][2*j+5]^255))))
                    else:
                        y.append(data[16][2*j+4]+(256*data[16][2*j+5]))
                #DNS        
                for i in range(18,33):
                    for j in range(1,14):
                        if(data[i][2*j+5]>=128):
                            y2.append(-1*((255^data[i][2*j+4])+1+(256*(data[i][2*j+5]^255))))
                        else:
                            y2.append(data[i][2*j+4]+(256*data[i][2*j+5]))
                for j in range(1,6):
                    if(data[33][2*j+5]>=128):
                        y2.append(-1*((255^data[16][2*j+4])+1+(256*(data[33][2*j+5]^255))))
                    else:
                        y2.append(data[33][2*j+4]+(256*data[33][2*j+5]))
            except:
                pass

            if(len(y2)==0 or len(y)==0):
                self.Button_ADC.setText("Try again")
            else:
                if(y2.index(max(y2))<20 or y.index(max(y))<20):
                    self.Button_ADC.setText("Try again")
                else:
                    self.ADC_Graph(x,y,y2)
                    self.Button_ADC.setText("ADC_Capture")
		
 
    def ADC_Graph(self,x,y,y2):
        try:
            self.fig.clear()
     
            ax = self.fig.add_subplot(111)
            ax.plot(x,y,'r')
            ax.plot(x,y2,'b')
            
            self.ADC.draw()

            now = datetime.now()

            self.Lbl_ADC_max_UPS.setText("["+str(now.hour)+":"+str(now.minute)+"]  UPS_max = [{0}, {1}]".format(y.index(max(y)),max(y)))
            self.Lbl_ADC_max_DNS.setText("DNS_max = [{0}, {1}]".format(y2.index(max(y2)),max(y2)))

            #self.Wave_capture()
        except:
            pass
            
if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MyWindow()
    window.show()
    app.exec_()
